"""
Excel Export Handler

This module handles exporting email data to Excel format using openpyxl.
"""

import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelHandler:
    """Handles exporting emails to Excel format."""

    def __init__(self):
        """Initialize Excel handler."""
        self.workbook = None
        self.worksheet = None

    def export_emails(
        self,
        emails: List[Dict[str, Any]],
        filename: str = None,
        output_dir: str = '.'
    ) -> str:
        """
        Export emails to Excel file.

        Args:
            emails: List of email dictionaries
            filename: Output filename (optional, auto-generated if not provided)
            output_dir: Output directory path

        Returns:
            Path to created Excel file

        Raises:
            Exception: If export fails
        """
        if not emails:
            raise ValueError("No emails to export")

        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"emails_{timestamp}.xlsx"

        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        # Create full path
        output_path = os.path.join(output_dir, filename)

        print(f"Exporting {len(emails)} emails to: {output_path}")

        try:
            # Create workbook and worksheet
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Emails"

            # Write data
            self._write_headers()
            self._write_email_data(emails)
            self._format_worksheet()

            # Save file
            self.workbook.save(output_path)
            print(f"Export successful! File saved to: {output_path}")

            return output_path

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            raise

    def _write_headers(self):
        """Write column headers."""
        headers = [
            'Date',
            'From',
            'To',
            'CC',
            'Subject',
            'Body Preview',
            'Labels',
            'Has Attachment',
            'Message ID',
            'Thread ID'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_email_data(self, emails: List[Dict[str, Any]]):
        """
        Write email data to worksheet.

        Args:
            emails: List of email dictionaries
        """
        for row, email in enumerate(emails, 2):  # Start from row 2 (after headers)
            # Truncate body for preview
            body_preview = email.get('body', '')[:200]
            if len(email.get('body', '')) > 200:
                body_preview += '...'

            # Write row data
            row_data = [
                email.get('date', ''),
                email.get('from', ''),
                email.get('to', ''),
                email.get('cc', ''),
                email.get('subject', ''),
                body_preview,
                email.get('labels', ''),
                'Yes' if email.get('has_attachment') else 'No',
                email.get('message_id', ''),
                email.get('thread_id', '')
            ]

            for col, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    def _format_worksheet(self):
        """Apply formatting to worksheet."""
        # Set column widths
        column_widths = {
            'A': 20,  # Date
            'B': 30,  # From
            'C': 30,  # To
            'D': 20,  # CC
            'E': 40,  # Subject
            'F': 60,  # Body Preview
            'G': 20,  # Labels
            'H': 15,  # Has Attachment
            'I': 30,  # Message ID
            'J': 30   # Thread ID
        }

        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width

        # Freeze header row
        self.worksheet.freeze_panes = 'A2'

        # Set row height for data rows
        for row in range(2, self.worksheet.max_row + 1):
            self.worksheet.row_dimensions[row].height = 40

    def export_emails_with_full_body(
        self,
        emails: List[Dict[str, Any]],
        filename: str = None,
        output_dir: str = '.'
    ) -> str:
        """
        Export emails with full body content (no truncation).

        Args:
            emails: List of email dictionaries
            filename: Output filename
            output_dir: Output directory path

        Returns:
            Path to created Excel file
        """
        if not emails:
            raise ValueError("No emails to export")

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"emails_full_{timestamp}.xlsx"

        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        output_path = os.path.join(output_dir, filename)

        print(f"Exporting {len(emails)} emails with full body to: {output_path}")

        try:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Emails"

            # Write headers (same as before)
            self._write_headers()

            # Write full email data
            for row, email in enumerate(emails, 2):
                row_data = [
                    email.get('date', ''),
                    email.get('from', ''),
                    email.get('to', ''),
                    email.get('cc', ''),
                    email.get('subject', ''),
                    email.get('body', ''),  # Full body, no truncation
                    email.get('labels', ''),
                    'Yes' if email.get('has_attachment') else 'No',
                    email.get('message_id', ''),
                    email.get('thread_id', '')
                ]

                for col, value in enumerate(row_data, 1):
                    cell = self.worksheet.cell(row=row, column=col)
                    cell.value = value
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Apply formatting
            self._format_worksheet()

            # Save file
            self.workbook.save(output_path)
            print(f"Export successful! File saved to: {output_path}")

            return output_path

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            raise


def main():
    """Test Excel export."""
    print("Excel Handler Test")
    print("-" * 50)

    # Sample data
    test_emails = [
        {
            'date': '2024-10-25 14:30:00',
            'from': 'sender@example.com',
            'to': 'recipient@example.com',
            'cc': '',
            'subject': 'Test Email 1',
            'body': 'This is a test email body with some content.',
            'labels': 'INBOX, IMPORTANT',
            'has_attachment': False,
            'message_id': 'msg123',
            'thread_id': 'thread123'
        },
        {
            'date': '2024-10-25 15:45:00',
            'from': 'another@example.com',
            'to': 'recipient@example.com',
            'cc': 'cc@example.com',
            'subject': 'Test Email 2 with Attachment',
            'body': 'This is another test email with a much longer body content to test text wrapping and formatting in Excel cells.',
            'labels': 'INBOX',
            'has_attachment': True,
            'message_id': 'msg456',
            'thread_id': 'thread456'
        }
    ]

    # Export test
    handler = ExcelHandler()
    output_path = handler.export_emails(
        test_emails,
        filename='test_emails.xlsx',
        output_dir='.'
    )

    print(f"\nTest export completed: {output_path}")


if __name__ == "__main__":
    main()
